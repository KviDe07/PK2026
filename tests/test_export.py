"""Тесты выгрузки сделок в Excel (export_deals)."""

from openpyxl import load_workbook

from admissions.export import _clean, export_deals


def test_clean_strips_bbcode_and_entities():
    assert _clean("[p]Привет[/p]") == "Привет"
    assert _clean("Re: тест &quot;кавычки&quot;") == 'Re: тест "кавычки"'
    assert _clean("") == "" and _clean(None) == ""


def test_export_builds_sheet_with_comments(cfg, fake_bitrix, tmp_path):
    contact = {"ID": "501", "LAST_NAME": "Иванов", "NAME": "Иван", "SECOND_NAME": "Иванович",
               "PHONE": [{"VALUE": "+79160001122"}], "EMAIL": [{"VALUE": "i@e.ru"}],
               "UF_CRM_PK_CODE": "111", "TYPE_ID": "CLIENT"}
    deal = {"ID": "9001", "CONTACT_ID": "501", "CATEGORY_ID": "8", "STAGE_ID": "C8:NEW",
            "TITLE": "Иванов — ТФ", "UF_CRM_B_CODE": "111", "UF_CRM_B_GROUP": "Техническая физика",
            "UF_CRM_B_SCORE": "287"}
    comments = {
        ("contact", "501"): [{"COMMENT": "[p]звонил, перезвонить[/p]", "AUTHOR_ID": "1", "CREATED": "2026-06-30T10:00:00+03:00"}],
        ("deal", "9001"): [{"COMMENT": "ждёт согласие", "AUTHOR_ID": "1", "CREATED": "2026-06-30T11:00:00+03:00"}],
    }
    client = fake_bitrix(contacts=[contact], deals=[deal], comments=comments)
    out = tmp_path / "export.xlsx"
    export_deals(cfg, out, client=client)

    wb = load_workbook(out)
    ws = wb["Сделки"]
    headers = [c.value for c in ws[1]]
    assert "ФИО" in headers and "Комментарии контакта" in headers and "Комментарии сделки" in headers
    row = {headers[i]: c.value for i, c in enumerate(ws[2])}
    assert row["ФИО"] == "Иванов Иван Иванович"
    assert row["Стадия"] == "Поступившие заявления"
    assert row["Тип"] == "Абитуриенты"
    assert row["Конкурсная группа"] == "Техническая физика"
    assert "звонил, перезвонить" in row["Комментарии контакта"]
    assert "Оператор Тест" in row["Комментарии контакта"]   # автор разрезолвлен
    assert "ждёт согласие" in row["Комментарии сделки"]
