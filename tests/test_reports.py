"""Smoke-тесты генерации Excel-отчётов и проверки документов."""

from openpyxl import load_workbook

from admissions.merge import process_records
from admissions.models import SOURCE_1C
from admissions.reports import build_reports, missing_documents


def _seed(db, cfg, make_record):
    recs = [
        make_record(
            SOURCE_1C, "A-1", status="enrolled",
            last_name="Иванов", first_name="Иван", birth_date="2007-03-15",
            passport="4509 1", snils="11223344595", education_doc="Аттестат", consent=True,
            program="Физика",
        ),
        make_record(
            SOURCE_1C, "A-2", status="admitted_to_exams",
            last_name="Кузнецова", first_name="Анна", birth_date="2007-09-30",
            snils="44455566677", program="Информатика",  # нет паспорта и аттестата
        ),
    ]
    process_records(db, cfg, recs, source=SOURCE_1C)


def test_missing_documents(db, cfg, make_record):
    _seed(db, cfg, make_record)
    ivanov = db.get_applicant(1)
    kuznetsova = db.get_applicant(2)
    assert missing_documents(cfg, ivanov) == []
    missing = missing_documents(cfg, kuznetsova)
    assert "Паспорт" in missing and "Документ об образовании" in missing


def test_build_reports_creates_all_sheets(db, cfg, make_record, tmp_path):
    _seed(db, cfg, make_record)
    out = build_reports(db, cfg, kinds=["all"], out_file=str(tmp_path / "r.xlsx"))
    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == ["Master", "Изменения", "Аналитика", "Review"]
    # в Master строка-заголовок + 2 абитуриента
    assert wb["Master"].max_row == 3


def test_build_single_sheet(db, cfg, make_record, tmp_path):
    _seed(db, cfg, make_record)
    out = build_reports(db, cfg, kinds=["analytics"], out_file=str(tmp_path / "a.xlsx"))
    wb = load_workbook(out)
    assert wb.sheetnames == ["Аналитика"]
