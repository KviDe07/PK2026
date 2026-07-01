"""Excel-отчёт «разбор сопоставления» по результату sync().

Листы:
  Тёзки         — несколько контактов с одним ФИО (различить нечем), пропущены;
  Конфликт ФИО  — ФИО совпало, но телефон другой → создан новый контакт, проверьте;
  Не создались  — контакт не записался (контроль дублей по телефону/почте);
  Выбывшие      — код есть на контакте, но абитуриент пропал из новой выгрузки;
  Отозванные заявления — сделка с кодом есть, но её заявления нет в новой выгрузке;
  Дубли         — несколько контактов с одним кодом / несколько сделок на группу.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Font


def write_problem_report(path: str | Path, stats: Dict[str, Any]) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    def add(title: str, headers, rows):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, 1):
            width = max([len(str(h))] + [len(str(r[i - 1])) for r in rows]) + 2
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width, 60)

    add("Тёзки", ["Код", "ФИО", "Причина"],
        [[a["code"], a["name"], a["reason"]] for a in stats.get("ambiguous", [])])
    add("Конфликт ФИО", ["Код", "ФИО", "Причина"],
        [[a["code"], a["name"], a["reason"]] for a in stats.get("conflicts", [])])
    add("Не создались", ["Код", "ФИО", "Причина"],
        [[a["code"], a["name"], a["reason"]] for a in stats.get("failed", [])])
    add("Выбывшие", ["Код", "ФИО", "ID контакта"],
        [[a["code"], a["name"], a["id"]] for a in stats.get("dropped", [])])
    add("Отозванные заявления", ["Код", "ID сделки", "Заявление", "Текущая стадия"],
        [[a["code"], a["deal"], a["key"], a["stage"]] for a in stats.get("withdrawn", [])])
    dups = ([["Дубль кода", c] for c in stats.get("code_dups", [])]
            + [["Дубль сделки", d] for d in stats.get("deal_dups", [])])
    add("Дубли", ["Тип", "Детали"], dups)

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
