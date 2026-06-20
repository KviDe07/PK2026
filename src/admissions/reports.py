"""Формирование Excel-отчётов и аналитики.

Листы:
  Master     — все абитуриенты с объединёнными полями и текущим статусом.
  Изменения  — история смен статуса и создания карточек.
  Аналитика  — сводки по статусам / направлениям / форме / основе + воронка.
  Review     — спорные слияния и расхождения полей (нужна ручная проверка).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import Config
from .db import Database
from .models import FIELD_LABELS
from .utils import timestamp_slug

# Заливка ячейки статуса по каноническому значению.
STATUS_FILLS = {
    "lead": "E7E6E6",
    "documents_submitted": "DDEBF7",
    "under_review": "FFF2CC",
    "admitted_to_exams": "FCE4D6",
    "passed": "D9E1F2",
    "consent_given": "E2EFDA",
    "enrolled": "C6E0B4",
    "rejected": "F8CBAD",
    "withdrawn": "D0CECE",
}

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=12)


# ── комплектность документов ──────────────────────────────────────────────────

def missing_documents(cfg: Config, row: Any) -> List[str]:
    """Список недостающих обязательных документов для карточки (или []).

    Проверяется только если статус достиг порога documents.required_from_status.
    """
    docs = cfg.settings.get("documents", {})
    required = docs.get("required_fields", [])
    from_status = docs.get("required_from_status")

    order = cfg.status_order
    if from_status in order:
        cur = row["current_status"]
        cur_rank = order.index(cur) if cur in order else -1
        if cur_rank < order.index(from_status):
            return []

    missing = []
    for fld in required:
        value = row[fld] if fld in row.keys() else None
        if value is None or (isinstance(value, str) and value.strip() == ""):
            missing.append(FIELD_LABELS.get(fld, fld))
    return missing


# ── листы ──────────────────────────────────────────────────────────────────────

_MASTER_COLUMNS = [
    "id", "last_name", "first_name", "middle_name", "birth_date",
    "program", "education_form", "funding_basis", "total_score",
    "current_status", "email", "phone", "snils", "consent",
    "onec_id", "bitrix_id", "updated_at",
]


def _sheet_master(wb: Workbook, db: Database, cfg: Config) -> None:
    ws = wb.create_sheet("Master")
    headers = [FIELD_LABELS.get(c, c) for c in _MASTER_COLUMNS] + ["Недостающие документы"]
    ws.append(headers)

    status_col = _MASTER_COLUMNS.index("current_status") + 1
    for a in db.all_applicants():
        row = [_cell(a, c) for c in _MASTER_COLUMNS]
        row.append(", ".join(missing_documents(cfg, a)))
        ws.append(row)
        # подсветка статуса
        status = a["current_status"]
        if status in STATUS_FILLS:
            ws.cell(row=ws.max_row, column=status_col).fill = PatternFill(
                "solid", fgColor=STATUS_FILLS[status]
            )
    _style_table(ws, len(headers))


def _sheet_changes(wb: Workbook, db: Database, cfg: Config) -> None:
    ws = wb.create_sheet("Изменения")
    ws.append(["Когда", "Абитуриент", "Тип", "Было", "Стало", "Источник", "Комментарий"])
    rows = db.query(
        "SELECT h.changed_at, h.previous_status, h.status, h.source, h.note, "
        "a.last_name, a.first_name, a.middle_name "
        "FROM status_history h JOIN applicants a ON a.id = h.applicant_id "
        "ORDER BY h.changed_at DESC, h.id DESC"
    )
    for r in rows:
        name = " ".join(p for p in (r["last_name"], r["first_name"], r["middle_name"]) if p)
        kind = "новая карточка" if not r["previous_status"] else "смена статуса"
        ws.append([
            r["changed_at"], name, kind,
            _status_label(cfg, r["previous_status"]), _status_label(cfg, r["status"]),
            r["source"], r["note"],
        ])
    _style_table(ws, 7)


def _sheet_analytics(wb: Workbook, db: Database, cfg: Config) -> None:
    ws = wb.create_sheet("Аналитика")
    total = db.query("SELECT COUNT(*) AS n FROM applicants")[0]["n"]

    ws.append(["Всего абитуриентов", total])
    ws.cell(row=1, column=1).font = _TITLE_FONT
    ws.append([])

    # Воронка по статусам (в порядке этапов)
    counts = {r["current_status"]: r["n"] for r in db.query(
        "SELECT current_status, COUNT(*) AS n FROM applicants GROUP BY current_status"
    )}
    _block(ws, "По статусам (воронка)", [
        (_status_label(cfg, s), counts.get(s, 0)) for s in cfg.status_order if counts.get(s, 0)
    ])
    # статусы вне списка порядка (на случай нестандартных)
    extra = [(s, n) for s, n in counts.items() if s not in cfg.status_order]
    if extra:
        _block(ws, "Прочие статусы", [(_status_label(cfg, s), n) for s, n in extra])

    _block(ws, "По направлениям", _group_counts(db, "program"))
    _block(ws, "По форме обучения", _group_counts(db, "education_form"))
    _block(ws, "По основе обучения", _group_counts(db, "funding_basis"))

    # Недостающие документы
    incomplete = sum(1 for a in db.all_applicants() if missing_documents(cfg, a))
    _block(ws, "Комплектность документов", [
        ("С недостающими документами", incomplete),
        ("Комплектно/не требуется", total - incomplete),
    ])

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14


def _sheet_review(wb: Workbook, db: Database, cfg: Config) -> None:
    ws = wb.create_sheet("Review")
    ws.append(["Тип", "Абитуриент/запись", "Источник", "Ключ", "Причина", "Кандидаты", "Score", "Когда"])
    kind_label = {"match": "неоднозначное слияние", "conflict": "расхождение полей"}
    for r in db.get_review_items(unresolved_only=True):
        ws.append([
            kind_label.get(r["kind"], r["kind"]), r["record_name"], r["source"],
            r["source_key"], r["reason"], r["candidate_ids"],
            r["score"], r["created_at"],
        ])
    if ws.max_row == 1:
        ws.append(["— спорных записей нет —"])
    _style_table(ws, 8)


# ── вспомогательное ──────────────────────────────────────────────────────────────

def _group_counts(db: Database, column: str) -> List[tuple]:
    rows = db.query(
        f"SELECT COALESCE({column}, '(не указано)') AS k, COUNT(*) AS n "
        f"FROM applicants GROUP BY k ORDER BY n DESC"
    )
    return [(r["k"], r["n"]) for r in rows]


def _block(ws: Worksheet, title: str, pairs: Sequence[tuple]) -> None:
    ws.append([title])
    ws.cell(row=ws.max_row, column=1).font = _TITLE_FONT
    for label, value in pairs:
        ws.append([label, value])
    ws.append([])


def _cell(row: Any, column: str) -> Any:
    value = row[column] if column in row.keys() else None
    if column == "consent":
        if value is None:
            return ""
        return "Да" if value else "Нет"
    if column == "current_status":
        return _STATUS_RU.get(value, value)
    return value


def _status_label(cfg: Config, status: Optional[str]) -> str:
    if not status:
        return ""
    # человекочитаемые названия статусов из status_mapping (обратное соответствие)
    return _STATUS_RU.get(status, status)


_STATUS_RU = {
    "lead": "Заявка/лид",
    "documents_submitted": "Документы поданы",
    "under_review": "На рассмотрении",
    "admitted_to_exams": "Допущен к экзаменам",
    "passed": "Прошёл по конкурсу",
    "consent_given": "Подал согласие",
    "enrolled": "Зачислен",
    "rejected": "Отказ/не прошёл",
    "withdrawn": "Забрал документы",
}


def _style_table(ws: Worksheet, n_cols: int) -> None:
    """Шапка, закрепление, автофильтр, ширины колонок."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{max(ws.max_row, 1)}"
    # авто-ширина (с ограничением)
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        width = 10
        for cell in ws[letter]:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 45))
        ws.column_dimensions[letter].width = width


def _expand_kinds(kinds: Sequence[str]) -> List[str]:
    if not kinds or "all" in kinds:
        return ["master", "changes", "analytics", "review"]
    return list(kinds)


def build_reports(
    db: Database, cfg: Config, kinds: Sequence[str] = ("all",), out_file: Optional[str] = None
) -> Path:
    """Собрать книгу Excel из запрошенных листов. Возвращает путь к файлу."""
    kinds = _expand_kinds(kinds)
    wb = Workbook()
    wb.remove(wb.active)  # убрать пустой лист по умолчанию

    builders = {
        "master": _sheet_master,
        "changes": _sheet_changes,
        "analytics": _sheet_analytics,
        "review": _sheet_review,
    }
    for kind in ["master", "changes", "analytics", "review"]:
        if kind in kinds:
            builders[kind](wb, db, cfg)

    if not wb.worksheets:
        _sheet_master(wb, db, cfg)

    if out_file:
        out_path = Path(out_file)
    else:
        cfg.output_dir.mkdir(parents=True, exist_ok=True)
        out_path = cfg.output_dir / f"report_{timestamp_slug()}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
