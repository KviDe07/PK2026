"""Симулятор зачисления бакалавриата МФТИ (прогноз проходных).

Перенос автономного скрипта sim.py в приложение: движок сохранён, но
  * КЦП берутся из config/kcp_bachelor.yaml (cfg.kcp / cfg.kcp_school);
  * шапка таблицы ищется автоматически по «Уникальный код» (не жёсткий header=5);
  * ошибки — RuntimeError (в вебе показываем сообщение, не роняем процесс);
  * функция возвращает путь к Excel и сводку (вместо print).

Логика зачисления (единая для всех конкурсных групп):
  1. Отдельная квота  — конкурс по приоритетам на квотных местах.
  2. Целевая квота    — конкурс целевиков.
  3. Особая квота     — конкурс льготников.
  4. ПЕРЕТОК: незанятые места всех квот добавляются к общему конкурсу.
  5. Общий конкурс: сначала БВИ (вне конкурса по баллам, между собой по ИД),
     затем остальные по СУММЕ баллов (предметы + ИД).
Модель приоритетов — deferred acceptance, глобально по всем группам до стабилизации.
"""

from __future__ import annotations

import logging
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .normalize import clean_str

log = logging.getLogger("admissions")

TYPE_RU = {'quota_otd': 'отд.квота', 'quota_cel': 'целевая',
           'quota_osob': 'особая', 'bvi': 'БВИ', 'general': 'общий'}

# Ожидаемые имена колонок выгрузки (левая часть — как в файле).
COL = {
    'group':    'Конкурсная группа',
    'bvi':      'Без вступительных испытаний',
    'consent':  'Согласие на зачисление',
    'priority': 'Приоритет',
    'sp':       'Сумма баллов по предметам',
    'si':       'Сумма баллов по ИД (все)',
    'ss':       'Сумма баллов',
    'code':     'Уникальный код',
    'fio':      'ФИО',
    'special':  'Особенности приема',
    'basis':    'Основание поступления',
    'control':  'Контроль пройден',
}

TICK = '✓'  # символ отметки в булевых колонках


# ============================================================================
#  ЗАГРУЗКА И ПОДГОТОВКА
# ============================================================================
def to_base_group(g, kcp_keys):
    """Свести целевую версию КГ ('... <заказчик>') к базовому названию из КЦП."""
    if not isinstance(g, str):
        return None
    for b in sorted(kcp_keys, key=len, reverse=True):
        if g == b or g.startswith(b + ' '):
            return b
    return g


def _read_table(path: str | Path) -> pd.DataFrame:
    """Прочитать выгрузку, найдя строку-шапку по колонке «Уникальный код»."""
    raw = pd.read_excel(path, sheet_name=0, header=None, dtype=object)
    header_row = None
    for i in range(min(30, len(raw))):
        values = [clean_str(v) for v in raw.iloc[i].tolist()]
        if COL['code'] in values:
            header_row = i
            break
    if header_row is None:
        raise RuntimeError("Не найдена шапка таблицы (нет колонки «Уникальный код»)")
    df = pd.read_excel(path, sheet_name=0, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    return df


def load(path: str | Path, kcp: Dict[str, tuple]) -> pd.DataFrame:
    p = Path(path)
    if not p.exists():
        raise RuntimeError(f"Файл не найден: {path}")
    try:
        df = _read_table(p)
    except RuntimeError:
        raise
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(f"Не удалось прочитать {path}: {e}")
    df = df.dropna(how='all').copy()

    missing = [v for v in COL.values() if v not in df.columns]
    if missing:
        raise RuntimeError("В выгрузке нет колонок: " + ", ".join(missing) +
                           ".\nДоступные: " + ", ".join(map(str, df.columns)))

    kcp_keys = list(kcp.keys())
    df['КГ'] = df[COL['group']].apply(lambda g: to_base_group(g, kcp_keys))
    df['_bvi'] = df[COL['bvi']] == TICK
    df['_consent'] = df[COL['consent']] == TICK
    df['_control'] = df[COL['control']] == TICK
    df['_prio'] = pd.to_numeric(df[COL['priority']], errors='coerce')
    df['_sp'] = pd.to_numeric(df[COL['sp']], errors='coerce').fillna(0)
    df['_si'] = pd.to_numeric(df[COL['si']], errors='coerce').fillna(0)
    df['_ss'] = pd.to_numeric(df[COL['ss']], errors='coerce').fillna(0)
    df['_code'] = df[COL['code']]
    return df


def build_apps(df: pd.DataFrame, kcp: Dict[str, tuple], require_control: bool = True):
    """Собрать заявления: код -> список dict с полями по каждой КГ.

    require_control=True: в конкурс идут только заявления с «Контроль пройден».
    Заявление без контроля выбывает, но абитуриент остаётся по остальным приоритетам.
    """
    bud = df[df[COL['basis']].isin(['Бюджетная основа', 'Целевой прием'])].copy()

    if require_control:
        before = len(bud)
        bud = bud[bud['_control']]
        log.info("Симулятор: фильтр «Контроль пройден» %d → %d заявлений (отсеяно %d)",
                 before, len(bud), before - len(bud))

    # баллы и БВИ берём из строк общих мест (там полный балл)
    score_map, bvi_map = {}, {}
    for _, r in bud[bud[COL['special']] == 'Общие места'].iterrows():
        k = (r['_code'], r['КГ'])
        score_map[k] = (r['_sp'], r['_si'], r['_ss'])
        if r['_bvi']:
            bvi_map[k] = True

    prio_map = {}
    for _, r in bud.iterrows():
        k = (r['_code'], r['КГ'])
        p = r['_prio']
        if pd.notna(p) and (k not in prio_map or p < prio_map[k]):
            prio_map[k] = int(p)

    otd_keys, cel_keys, osob_keys = set(), set(), set()
    for _, r in bud.iterrows():
        k = (r['_code'], r['КГ'])
        if r[COL['special']] == 'Отдельная квота':
            otd_keys.add(k)
        if r[COL['basis']] == 'Целевой прием':
            cel_keys.add(k)
        if isinstance(r[COL['special']], str) and 'Особая' in r[COL['special']]:
            osob_keys.add(k)

    consent = df.groupby('_code')['_consent'].max().to_dict()
    fio = df.groupby('_code')[COL['fio']].first().to_dict()

    apps: Dict[Any, list] = {}
    for k in (set(score_map) | otd_keys | cel_keys | osob_keys):
        code, kg = k
        if kg not in kcp:
            continue
        sp, si, ss = score_map.get(k, (float('nan'),) * 3)
        if pd.isna(sp):  # только квотная строка без общей — тянем баллы откуда есть
            rr = bud[(bud['_code'] == code) & (bud['КГ'] == kg) & (bud['_sp'] > 0)]
            if len(rr):
                sp, si, ss = rr['_sp'].iloc[0], rr['_si'].iloc[0], rr['_ss'].iloc[0]
            else:
                sp, si, ss = 0, 0, 0
        apps.setdefault(code, []).append({
            'kg': kg, 'prio': prio_map.get(k, 999),
            'sp': float(sp), 'si': float(si), 'ss': float(ss),
            'bvi': bvi_map.get(k, False),
            'in_general': k in score_map,
            'otd': k in otd_keys, 'cel': k in cel_keys, 'osob': k in osob_keys,
        })
    for c in apps:
        apps[c].sort(key=lambda a: a['prio'])
    return apps, consent, fio


# ============================================================================
#  ДВИЖОК ЗАЧИСЛЕНИЯ
# ============================================================================
def rank_key(a):
    # БВИ вне конкурса по баллам: всегда выше. Среди БВИ — по ИД, затем сумме.
    # Обычные — по сумме баллов, затем по предметам.
    if a['bvi']:
        return (0, -a['si'], -a['ss'])
    return (1, -a['ss'], -a['sp'])


def run_priority(apps, pool, cap, field_filter):
    """Deferred acceptance на местах cap[kg] среди заявлений, прошедших field_filter."""
    cand = {}
    for c in pool:
        lst = [a for a in apps[c] if field_filter(a)]
        if lst:
            cand[c] = sorted(lst, key=lambda a: a['prio'])
    ptr = {c: 0 for c in cand}
    groups = {kg: [] for kg in cap}
    active = set(cand)
    while active:
        nxt = set()
        prop = {}
        for c in list(active):
            i = ptr[c]
            lst = cand[c]
            if i >= len(lst):
                continue
            prop.setdefault(lst[i]['kg'], []).append(c)
        for kg, cs in prop.items():
            groups[kg].extend(cs)
        for kg in set(prop):
            def sc(c):
                a = next(x for x in cand[c] if x['kg'] == kg)
                return rank_key(a)
            ms = sorted(set(groups[kg]), key=sc)
            groups[kg] = ms[:cap[kg]]
            for c in ms[cap[kg]:]:
                ptr[c] += 1
                nxt.add(c)
        active = nxt
    placed = {}
    for kg, cs in groups.items():
        for c in cs:
            placed[c] = kg
    return placed


def simulate(apps, consent, kcp: Dict[str, tuple], only_consent: bool = True):
    pool = [c for c in apps if (consent.get(c) or not only_consent)]
    assigned = {}

    # 1. отдельная квота
    otd = run_priority(apps, pool, {kg: kcp[kg][2] for kg in kcp}, lambda a: a['otd'])
    for c, kg in otd.items():
        assigned[c] = {'kg': kg, 'type': 'quota_otd'}

    # 2. целевая квота
    rem = [c for c in pool if c not in assigned]
    cel = run_priority(apps, rem, {kg: kcp[kg][3] for kg in kcp}, lambda a: a['cel'])
    for c, kg in cel.items():
        assigned[c] = {'kg': kg, 'type': 'quota_cel'}

    # 3. особая квота
    rem = [c for c in pool if c not in assigned]
    osob = run_priority(apps, rem, {kg: kcp[kg][1] for kg in kcp}, lambda a: a['osob'])
    for c, kg in osob.items():
        assigned[c] = {'kg': kg, 'type': 'quota_osob'}

    used = {
        'otd':  Counter(v['kg'] for v in assigned.values() if v['type'] == 'quota_otd'),
        'cel':  Counter(v['kg'] for v in assigned.values() if v['type'] == 'quota_cel'),
        'osob': Counter(v['kg'] for v in assigned.values() if v['type'] == 'quota_osob'),
    }

    # 4. общий конкурс = основной общий + ВСЕ незанятые квотные места
    gen_cap = {}
    for kg in kcp:
        gen_cap[kg] = (kcp[kg][0]
                       + (kcp[kg][1] - used['osob'].get(kg, 0))
                       + (kcp[kg][2] - used['otd'].get(kg, 0))
                       + (kcp[kg][3] - used['cel'].get(kg, 0)))
    rem = [c for c in pool if c not in assigned]
    gen = run_priority(apps, rem, gen_cap, lambda a: a['in_general'])
    for c, kg in gen.items():
        a = next(x for x in apps[c] if x['kg'] == kg and x['in_general'])
        assigned[c] = {'kg': kg, 'type': 'bvi' if a['bvi'] else 'general'}

    return assigned, gen_cap, used


# ============================================================================
#  ТАБЛИЦЫ
# ============================================================================
def make_summary(apps, consent, res, gen_cap, used, kcp, school, only_consent):
    rows = []
    for kg in kcp:
        v = kcp[kg]
        if sum(v) == 0:
            continue
        mem = [(c, x['type']) for c, x in res.items() if x['kg'] == kg]
        tc = Counter(t for _, t in mem)
        gen_sums = [next(x for x in apps[c] if x['kg'] == kg and x['in_general'])['ss']
                    for c, t in mem if t == 'general']
        prohodnoy = int(min(gen_sums)) if gen_sums else '—'
        applied = sum(1 for c, al in apps.items()
                      if (consent.get(c) or not only_consent) and any(x['kg'] == kg for x in al))
        rows.append({
            'Физтех-школа': school.get(kg, ''),
            'Конкурсная группа': kg,
            'КЦП всего': sum(v), 'Общий пул': gen_cap[kg], 'Зачислено': len(mem),
            'БВИ': tc.get('bvi', 0), 'Общий конкурс': tc.get('general', 0),
            'Отд.квота': tc.get('quota_otd', 0), 'Целевая': tc.get('quota_cel', 0),
            'Проходной (сумма)': prohodnoy,
            'Заполнено %': round(100 * len(mem) / sum(v)) if sum(v) else 0,
            'Подали': applied, 'Недобор': max(0, sum(v) - len(mem)),
        })
    summ = pd.DataFrame(rows).sort_values(['Физтех-школа', 'Конкурсная группа']).reset_index(drop=True)
    summ.insert(0, '№', range(1, len(summ) + 1))
    return summ


def make_group_table(apps, consent, res, kgt, only_consent):
    final_kg = {c: v['kg'] for c, v in res.items()}
    final_type = {c: v['type'] for c, v in res.items()}
    rows = []
    for code, al in apps.items():
        if only_consent and not consent.get(code):
            continue
        a = next((x for x in al if x['kg'] == kgt and x['in_general']), None) \
            or next((x for x in al if x['kg'] == kgt), None)
        if a is None:
            continue
        rows.append({
            'ФИО': None,  # заполним при выводе
            '_code': code,
            'Балл предметы': a['sp'], 'ИД': a['si'], 'Сумма': a['ss'],
            'БВИ': 'БВИ' if a['bvi'] else '', 'Приоритет сюда': a['prio'],
            'Проходит': final_kg.get(code) == kgt,
            'Итог: зачислен на': final_kg.get(code, '— не проходит —'),
            'Тип места': TYPE_RU.get(final_type.get(code, ''), ''),
        })
    d = pd.DataFrame(rows)
    if len(d) == 0:
        return d
    d['_b'] = d['БВИ'] == 'БВИ'
    d = d.sort_values(['_b', 'Сумма', 'Балл предметы'],
                      ascending=[False, False, False]).drop(columns='_b').reset_index(drop=True)
    return d


# ============================================================================
#  EXCEL-ОФОРМЛЕНИЕ
# ============================================================================
NAVY = '1F4E78'; GREEN_F = 'C6EFCE'; GREEN_T = '006100'; AMBER = 'FFF2CC'
GREY = '555555'; LIGHT = 'EEF3F8'


def _f(**k):
    return Font(name='Arial', **k)


def _sheet_name(kg, used_names):
    base = kg[:28]
    n, i = base, 2
    while n in used_names:
        n = base[:26] + f'_{i}'
        i += 1
    used_names.add(n)
    return n


def write_excel(out, summ, tables, res, gen_cap, used, kcp, school, only_consent, fio):
    HEAD_FILL = PatternFill('solid', fgColor=NAVY)
    HEAD_FONT = _f(size=10, bold=True, color='FFFFFF')
    thin = Side(style='thin', color='D0D0D0')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    C = Alignment(horizontal='center', vertical='center')
    L = Alignment(horizontal='left', vertical='center')
    GREEN_FILL = PatternFill('solid', fgColor=GREEN_F)
    AMBER_FILL = PatternFill('solid', fgColor=AMBER)
    ZEBRA = PatternFill('solid', fgColor=LIGHT)

    wb = Workbook()

    # ---------- СВОДКА ----------
    ws = wb.active
    ws.title = 'Сводка'
    scope = 'все абитуриенты' if not only_consent else 'только с согласием на зачисление'
    ws.merge_cells('A1:N1')
    ws['A1'] = 'Симуляция зачисления МФТИ — сводка по всем направлениям'
    ws['A1'].font = _f(size=14, bold=True, color=NAVY); ws['A1'].alignment = L
    tot_kcp = sum(sum(kcp[k]) for k in kcp)
    ws.merge_cells('A2:N2')
    ws['A2'] = (f'Учтены: {scope}.  Всего бюджетных мест: {tot_kcp}.  '
                f'Зачислено (проходят сейчас): {int(summ["Зачислено"].sum())}.')
    ws['A2'].font = _f(size=10, italic=True, color=GREY); ws['A2'].alignment = L
    ws.merge_cells('A3:N3')
    ws['A3'] = ('Логика: квоты → незанятые места квот перетекают в общий конкурс → БВИ (вне конкурса) + '
                'общий по сумме баллов. Приоритеты учтены глобально. '
                'Проходной — минимальная сумма баллов среди прошедших общим конкурсом (без БВИ).')
    ws['A3'].font = _f(size=9, italic=True, color=GREY); ws['A3'].alignment = L

    cols = list(summ.columns)
    widths = {'№': 5, 'Физтех-школа': 11, 'Конкурсная группа': 50, 'КЦП всего': 9,
              'Общий пул': 10, 'Зачислено': 10, 'БВИ': 6, 'Общий конкурс': 13,
              'Отд.квота': 10, 'Целевая': 9, 'Проходной (сумма)': 16, 'Заполнено %': 12,
              'Подали': 9, 'Недобор': 9}
    hrow = 5
    for j, cn in enumerate(cols, 1):
        c = ws.cell(hrow, j, cn)
        c.font = HEAD_FONT; c.fill = HEAD_FILL; c.alignment = C; c.border = BORDER
        ws.column_dimensions[get_column_letter(j)].width = widths.get(cn, 11)
    for i, (_, row) in enumerate(summ.iterrows()):
        r = hrow + 1 + i
        for j, cn in enumerate(cols, 1):
            c = ws.cell(r, j, row[cn]); c.border = BORDER
            c.alignment = L if cn in ('Конкурсная группа', 'Физтех-школа') else C
            c.font = _f(size=10)
            if i % 2 == 1:
                c.fill = ZEBRA
            if cn == 'Заполнено %' and row[cn] >= 100:
                c.fill = GREEN_FILL; c.font = _f(size=10, color=GREEN_T, bold=True)
            if cn == 'Проходной (сумма)' and row['Заполнено %'] < 100:
                c.font = _f(size=10, color=GREY, italic=True)
    ws.freeze_panes = 'A6'
    nr = hrow + 1 + len(summ) + 1
    ws.merge_cells(f'A{nr}:N{nr}')
    ws.cell(nr, 1, 'Примечание: где заполнено <100%, проходной балл не показателен — '
                   'конкурса нет, проходят все подавшие (низкая сумма = олимпиадник/БВИ без баллов ЕГЭ).'
            ).font = _f(size=9, italic=True, color=GREY)

    # ---------- ЛИСТЫ ПО ГРУППАМ (ФАКТ первым) ----------
    GCOLS = [('№', '№', 6), ('ФИО', 'ФИО', 30), ('Балл предметы', 'Балл ЕГЭ', 12),
             ('ИД', 'ИД', 7), ('Сумма', 'Сумма', 9), ('БВИ', 'БВИ', 7),
             ('Приоритет сюда', 'Приор.', 9), ('Итог: зачислен на', 'Итог: зачислен на', 42),
             ('Тип места', 'Тип', 11)]
    used_names = {'Сводка'}
    order = sorted(tables.keys(),
                   key=lambda k: (school.get(k, '') != 'ФАКТ', school.get(k, ''), k))
    for kg in order:
        df = tables[kg]
        v = kcp[kg]
        ws = wb.create_sheet(_sheet_name(kg, used_names))
        passed = int(df['Проходит'].sum()) if len(df) else 0
        ws.merge_cells('A1:I1')
        ws['A1'] = f'{school.get(kg, "")} — {kg}'
        ws['A1'].font = _f(size=12, bold=True, color=NAVY); ws['A1'].alignment = L
        ws.merge_cells('A2:I2')
        ws['A2'] = (f'КЦП всего {sum(v)}: общий {v[0]}, особая {v[1]}, отдельная {v[2]}, целевая {v[3]}.  '
                    f'Занято квот: отд {used["otd"].get(kg, 0)}/цел {used["cel"].get(kg, 0)}/особ {used["osob"].get(kg, 0)}.  '
                    f'Общий пул (с перетоком) = {gen_cap[kg]}.')
        ws['A2'].font = _f(size=9, italic=True, color=GREY); ws['A2'].alignment = L
        ws.merge_cells('A3:I3')
        fill_pct = round(100 * passed / sum(v)) if sum(v) else 0
        ws['A3'] = f'Подали: {len(df)}.  Проходят сейчас (зелёные): {passed}.  Заполнено: {fill_pct}%.'
        ws['A3'].font = _f(size=9, italic=True, color=GREY); ws['A3'].alignment = L
        hrow = 5
        for j, (k, t, w) in enumerate(GCOLS, 1):
            c = ws.cell(hrow, j, t)
            c.font = HEAD_FONT; c.fill = HEAD_FILL; c.alignment = C; c.border = BORDER
            ws.column_dimensions[get_column_letter(j)].width = w
        for i, (_, row) in enumerate(df.iterrows()):
            r = hrow + 1 + i
            passes = bool(row['Проходит'])
            isb = row['БВИ'] == 'БВИ'
            values = {'№': i + 1, 'ФИО': fio.get(row['_code'], ''),
                      'Балл предметы': int(row['Балл предметы']), 'ИД': int(row['ИД']),
                      'Сумма': int(row['Сумма']), 'БВИ': row['БВИ'],
                      'Приоритет сюда': row['Приоритет сюда'],
                      'Итог: зачислен на': row['Итог: зачислен на'], 'Тип места': row['Тип места']}
            for j, (k, t, w) in enumerate(GCOLS, 1):
                c = ws.cell(r, j, values[k]); c.border = BORDER
                c.alignment = L if k in ('ФИО', 'Итог: зачислен на') else C
                if passes:
                    c.fill = GREEN_FILL; c.font = _f(size=10, color=GREEN_T)
                elif isb:
                    c.fill = AMBER_FILL; c.font = _f(size=10)
                else:
                    c.font = _f(size=10)
        ws.freeze_panes = 'A6'

    # ---------- МЕТОДИКА ----------
    ws = wb.create_sheet('Методика')
    notes = [
        'МЕТОДИКА СИМУЛЯЦИИ ЗАЧИСЛЕНИЯ (МФТИ)', '',
        f'Режим: {"все абитуриенты" if not only_consent else "только с согласием на зачисление"}.',
        'КЦП берутся из config/kcp_bachelor.yaml (источник pk.mipt.ru/bachelor/2026_places).', '',
        'ПОРЯДОК ЗАЧИСЛЕНИЯ (единый для всех конкурсных групп):',
        '  1. Отдельная квота — конкурс по приоритетам на квотных местах.',
        '  2. Целевая квота — конкурс целевиков.',
        '  3. Особая квота — конкурс льготников.',
        '  4. ПЕРЕТОК: незанятые места всех квот добавляются к общему конкурсу.',
        '  5. Общий конкурс: сначала БВИ (вне конкурса по баллам, между собой по ИД),',
        '     затем остальные по СУММЕ баллов (предметы + ИД).', '',
        'МОДЕЛЬ ПРИОРИТЕТОВ (классическая, deferred acceptance):',
        '  • Согласие = согласие на зачисление в МФТИ (не на конкретную группу).',
        '  • Абитуриент прогоняется по всем приоритетам сверху вниз, занимает место на',
        '    ВЫСШЕМ проходящем приоритете; из нижних выбывает (отток вверх).',
        '  • Конкурс глобальный по всем группам, итерации до стабилизации.', '',
        'ПРОХОДНОЙ БАЛЛ: минимальная сумма баллов (предметы + ИД) среди зачисленных',
        'ОБЩИМ конкурсом (БВИ и квотники не учитываются).',
        'Где заполнено <100% — проходной не показателен, конкурса нет.', '',
        'ДОПУЩЕНИЯ:',
        '  • Незанятые места особой/отдельной/целевой квот перетекают в общий конкурс.',
        '  • Шкала «баллов за предметы» различается между направлениями (число ЕГЭ/ДВИ),',
        '    поэтому проходной сравним только ВНУТРИ направления.',
        '  • Тайбрейк при равных: сумма → предметы.',
    ]
    for i, l in enumerate(notes, 1):
        c = ws.cell(i, 1, l)
        if i == 1:
            c.font = _f(size=12, bold=True, color=NAVY)
        elif l.isupper() or l.endswith(':'):
            c.font = _f(size=10, bold=True)
        else:
            c.font = _f(size=10)
    ws.column_dimensions['A'].width = 92

    out = Path(out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


# ============================================================================
#  ОРКЕСТРАТОР
# ============================================================================
def simulate_admission(cfg, path: str | Path, out_path: str | Path,
                       only_consent: bool = True, require_control: bool = True) -> Dict[str, Any]:
    """Прогнать симуляцию и записать Excel. Возвращает сводку (числа + путь)."""
    kcp = cfg.kcp
    school = cfg.kcp_school
    if not kcp:
        raise RuntimeError("КЦП не заданы — заполните config/kcp_bachelor.yaml")

    df = load(path, kcp)
    apps, consent, fio = build_apps(df, kcp, require_control=require_control)
    res, gen_cap, used = simulate(apps, consent, kcp, only_consent=only_consent)

    summ = make_summary(apps, consent, res, gen_cap, used, kcp, school, only_consent)
    tables = {kg: make_group_table(apps, consent, res, kg, only_consent)
              for kg in kcp if sum(kcp[kg]) > 0}
    tables = {kg: t for kg, t in tables.items() if len(t) > 0}

    out = write_excel(out_path, summ, tables, res, gen_cap, used, kcp, school, only_consent, fio)
    return {
        "path": out,
        "applicants": len(apps),
        "with_consent": sum(1 for c in apps if consent.get(c)),
        "enrolled": len(res),
        "groups": len(tables),
        "only_consent": only_consent,
        "require_control": require_control,
    }
