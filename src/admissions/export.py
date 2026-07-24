"""Выгрузка сделок воронки бакалавриата в Excel (с комментариями).

По каждой сделке: контакт (ФИО/тел/почта/код/тип), данные заявления (Б:), стадия,
а также таймлайн-заметки КОНТАКТА и комментарии по СДЕЛКЕ (из ленты Битрикса).
"""

from __future__ import annotations

import html
import logging
import re
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .bitrix_client import BitrixClient, extract_value
from .bitrix_fields import index_by_xml

log = logging.getLogger("admissions")

_BB = re.compile(r"\[/?[^\]]+\]")  # грубая чистка BBCode ([p], [b], [URL=...], ...)

# Колонки данных сделки в выгрузке, по уровню. Ключ — XML_ID нашего поля (B_/M_/A_)
# ИЛИ прямой код готового поля портала (UF_CRM_…). enum-поля (Кафедра/УГС/Специальность)
# резолвятся из ID в текст. Готовые поля есть на портале, но без XML_ID.
_INTERVIEW_COLS = [
    ("UF_CRM_1752747663", "Группа собеседование"),
    ("UF_CRM_1752747696", "Комментарий собеседование"),
]
_EXPORT_COLS = {
    "bachelor": [
        ("B_GROUP", "Конкурсная группа"), ("B_SCORE", "Сумма баллов"),
        ("B_PRIORITY", "Приоритет"), ("B_SCORE_ID", "Сумма баллов по ИД"),
        ("B_CONSENT", "Согласие"), ("B_BVI", "Без ВИ"), ("B_BASIS", "Основание"),
        ("B_TARGETED", "Целевик"), ("B_SPECIAL", "Особое право"),
        ("B_APP_DATE", "Дата подачи"), ("B_CONTROL", "Контроль"),
    ] + _INTERVIEW_COLS,
    "master": [
        ("M_GROUP", "Конкурсная группа"), ("M_SCORE", "Сумма баллов"),
        ("M_PRIORITY", "Приоритет"), ("M_SCORE_ID", "Сумма баллов по ИД"),
        ("M_CONSENT", "Согласие"), ("M_BASIS", "Основание"), ("M_FEATURES", "Особенности"),
        ("UF_CRM_1750780247059", "Кафедра"),
    ] + _INTERVIEW_COLS,
    "postgrad": [
        ("A_GROUP", "Конкурсная группа"), ("A_SCORE", "Сумма баллов"),
        ("A_PRIORITY", "Приоритет"), ("A_SCORE_ID", "Сумма баллов по ИД"),
        ("A_CONSENT", "Согласие"), ("A_BASIS", "Основание"), ("A_FEATURES", "Особенности"),
        ("A_APP_DATE", "Дата подачи"),
        ("UF_CRM_1750624562799", "Укрупнённая группа специальностей"),
        ("UF_CRM_1750624913519", "Специальность"),
    ] + _INTERVIEW_COLS,
}


def _clean(text: Any) -> str:
    if not text:
        return ""
    return _BB.sub("", html.unescape(str(text))).strip()


def _fetch_comments(client: BitrixClient, entity_type: str, ids: List[str]) -> Dict[str, list]:
    """{entity_id: [сырые комментарии ленты]} пачками."""
    out: Dict[str, list] = {}
    ops = [("crm.timeline.comment.list",
            {"filter": {"ENTITY_TYPE": entity_type, "ENTITY_ID": i},
             "select": ["COMMENT", "AUTHOR_ID", "CREATED"],
             "order": {"CREATED": "ASC"}}) for i in ids]
    for i, r in zip(ids, client.batch_call(ops)):
        out[str(i)] = r.get("result") or []
    return out


def _resolve_users(client: BitrixClient, author_ids) -> Dict[str, str]:
    """Имена авторов комментариев (user.get). Требует scope `user` у вебхука;
    если прав нет (CRM-only вебхук) — тихо возвращаем пусто, автор не показывается."""
    ids = sorted({str(a) for a in author_ids if a})
    users: Dict[str, str] = {}
    if not ids:
        return users
    for i, r in zip(ids, client.batch_call([("user.get", {"ID": i}) for i in ids])):
        if r.get("error"):
            continue
        res = r.get("result")
        u = res[0] if isinstance(res, list) and res else (res if isinstance(res, dict) else None)
        if u:
            users[i] = " ".join(x for x in [u.get("LAST_NAME"), u.get("NAME")] if x) or u.get("EMAIL", "")
    return users


def _format(raw: list, users: Dict[str, str]) -> str:
    parts = []
    for cm in raw:
        txt = _clean(cm.get("COMMENT"))
        if not txt:
            continue
        date = (cm.get("CREATED") or "")[:10]
        who = users.get(str(cm.get("AUTHOR_ID")), "")
        prefix = f"[{date}] {who}: " if who else f"[{date}] "
        parts.append(prefix + txt)
    return "\n".join(parts)


def export_deals(cfg, path: str | Path, client=None, level: str = "bachelor") -> Path:
    client = client or BitrixClient.from_config(cfg)
    cat = cfg.category_id_for(level)

    dcode = {x: r["FIELD_NAME"] for x, r in index_by_xml(client, "deal").items()}
    cidx = index_by_xml(client, "contact")
    ccode = cidx["PK_CODE"]["FIELD_NAME"]

    # колонки данных по уровню → (FIELD_NAME, заголовок, enum_map|None).
    # Ключ-XML_ID (B_/M_/A_) резолвим через dcode; прямой код готового поля берём как есть.
    # enum-поля (Кафедра/УГС/Специальность) резолвим из ID в текст по crm.deal.fields.
    meta = client._call("crm.deal.fields", {}).get("result", {}) or {}
    export_cols = []
    for key, title in _EXPORT_COLS.get(level, _EXPORT_COLS["bachelor"]):
        fname = dcode.get(key, key)
        m = meta.get(fname) or {}
        emap = ({str(it["ID"]): it["VALUE"] for it in (m.get("items") or [])}
                if m.get("type") == "enumeration" else None)
        export_cols.append((fname, title, emap))
    known = set(dcode.values())
    extra_fields = [fn for fn, _, _ in export_cols if fn not in known]

    stages = client._call("crm.status.list",
                          {"filter": {"ENTITY_ID": f"DEAL_STAGE_{cat}" if cat else "DEAL_STAGE"}}
                          ).get("result", []) or []
    stage_name = {s["STATUS_ID"]: s["NAME"] for s in stages}
    ctypes = client._call("crm.status.list", {"filter": {"ENTITY_ID": "CONTACT_TYPE"}}).get("result", []) or []
    type_label = {s["STATUS_ID"]: s["NAME"] for s in ctypes}  # CLIENT -> «Абитуриенты»

    deals = client.list_all("crm.deal.list", filter={"CATEGORY_ID": cat},
                            select=["ID", "TITLE", "STAGE_ID", "CONTACT_ID"]
                                   + list(dcode.values()) + extra_fields)

    cids = sorted({str(d["CONTACT_ID"]) for d in deals
                   if d.get("CONTACT_ID") and str(d["CONTACT_ID"]) != "0"})
    contacts: Dict[str, Dict[str, Any]] = {}
    for i in range(0, len(cids), 50):
        for x in client.list_all("crm.contact.list", filter={"ID": cids[i:i + 50]},
                                  select=["ID", "LAST_NAME", "NAME", "SECOND_NAME",
                                          "PHONE", "EMAIL", "TYPE_ID", ccode]):
            contacts[str(x["ID"])] = x

    c_comments = _fetch_comments(client, "contact", cids)
    d_comments = _fetch_comments(client, "deal", [str(d["ID"]) for d in deals])
    authors = [cm.get("AUTHOR_ID") for lst in list(c_comments.values()) + list(d_comments.values())
               for cm in lst]
    users = _resolve_users(client, authors)

    wb = Workbook()
    ws = wb.active
    ws.title = "Сделки"
    headers = (["ID сделки", "Стадия", "ФИО", "Телефон", "Почта", "Уникальный код", "Тип"]
               + [t for _, t, _ in export_cols]
               + ["Комментарии контакта", "Комментарии сделки"])
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for d in deals:
        cid = str(d.get("CONTACT_ID") or "")
        c = contacts.get(cid, {})
        fio = " ".join(x for x in [c.get("LAST_NAME"), c.get("NAME"), c.get("SECOND_NAME")] if x)
        row = [
            d.get("ID"), stage_name.get(d.get("STAGE_ID"), d.get("STAGE_ID")), fio,
            extract_value(c.get("PHONE")), extract_value(c.get("EMAIL")),
            c.get(ccode), type_label.get(c.get("TYPE_ID"), c.get("TYPE_ID")),
        ]
        for fname, _, emap in export_cols:
            val = d.get(fname)
            if emap and val not in (None, ""):
                val = emap.get(str(val), val)   # enum: ID → текст
            row.append(val)
        row += [_format(c_comments.get(cid, []), users),
                _format(d_comments.get(str(d["ID"]), []), users)]
        ws.append(row)

    # ширины + перенос для колонок комментариев
    widths = [10, 22, 28, 16, 26, 14, 12] + [16] * len(export_cols) + [50, 50]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w
    last = len(headers)
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=last - 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=last).alignment = Alignment(wrap_text=True, vertical="top")

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.info("Выгружено сделок: %d", len(deals))
    return path
